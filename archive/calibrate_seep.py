"""
SEEP/W Hydraulic Calibration via GSI
=====================================
Calibrates Ksat and anisotropy ratio against observed volumetric water
content from the S2 moisture sensors directly using the GeoStudio Python API.

Parameters:
  - KSat_WYC  : saturated hydraulic conductivity of Seep WYC (ft/s)
  - KSat_AWYC : scale factor applied to Seep AWYC KYXRatio
  - KSat_UYC  : saturated hydraulic conductivity of Seep UYC (ft/s)

Calibration target:
  - Minimise RMSE between simulated and observed VWC at 1.5m and 3m depth.
  - Calibration window: 2018-10-18 to 2018-12-16 (Requested)
"""

import sys, os, shutil, warnings
import numpy as np
import pandas as pd
from scipy.optimize import minimize

try:
    import grpc
    from google.protobuf.json_format import MessageToDict
    from google.protobuf.struct_pb2 import Value, ListValue, Struct
    import gsi
except ImportError as e:
    print(f"GSI import error: {e}")
    sys.exit(1)

warnings.filterwarnings("ignore")

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------
CALIB_DIR       = r"E:\Github\MCP_Geostudio\calibration"
BASE_GSZ        = os.path.join(CALIB_DIR, "Metro-Center_cal.gsz")
WORK_GSZ        = os.path.join(CALIB_DIR, "_seep_incl_work.gsz")
OUT_GSZ         = os.path.join(CALIB_DIR, "Metro-Center-seep-final.gsz")
SENSOR_CSV      = os.path.join(CALIB_DIR, "S2.csv")
OUT_CSV         = os.path.join(CALIB_DIR, "seep_calibration_trial_log.csv")

# NEW Requested Calibration event window
CALIB_START     = "2018-10-18"
CALIB_END       = "2018-12-16"

CREST_X         = 195.0
SURFACE_Y       = 83.0      # surface elev (ft)
DEPTH_1_5M      = 4.92      # ft
DEPTH_3M        = 9.84      # ft

SENSOR_Y_SHALLOW = SURFACE_Y - DEPTH_1_5M 
SENSOR_Y_DEEP    = SURFACE_Y - DEPTH_3M   

SEEP_INITIAL    = "Initial Condition"
SEEP_TRANSIENT  = "Rainfall Simulation"

KSAT_WYC_BASE        = 1.004e-07   
KYX_RATIO_AWYC_BASE  = 11155.0     
KSAT_UYC_BASE        = 1.004e-07   

LOG_KSAT_WYC_MIN     = np.log10(KSAT_WYC_BASE) - 2.0
LOG_KSAT_WYC_MAX     = np.log10(KSAT_WYC_BASE) + 2.0
LOG_KYX_MIN          = np.log10(1000)
LOG_KYX_MAX          = np.log10(500000)
LOG_KSAT_UYC_MIN     = np.log10(KSAT_UYC_BASE) - 2.0
LOG_KSAT_UYC_MAX     = np.log10(KSAT_UYC_BASE) + 2.0

MAX_OPT_ITER    = 100

# Manual node index override - set these if find_sensor_nodes() fails.
# Run once with None to auto-detect; copy the printed idx values here afterwards.
SENSOR_NODE_SHALLOW = None   # e.g. 1042
SENSOR_NODE_DEEP    = None   # e.g. 987

# ---------------------------------------------------------------------------
# Load Sensor Data
# ---------------------------------------------------------------------------
def load_sensor_data():
    df = pd.read_csv(SENSOR_CSV, low_memory=False)
    df["timestamp"] = pd.to_datetime(df["timestamp"], errors="coerce")
    df = df.dropna(subset=["timestamp"])
    for col in ["Moisture_1.5m", "Moisture_3m", "Precipitation"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    crest = df[df["position"] == "Crest"].copy().set_index("timestamp")
    daily = crest[["Moisture_1.5m", "Moisture_3m", "Precipitation"]].resample("D").agg({
        "Moisture_1.5m": "mean",
        "Moisture_3m":   "mean",
        "Precipitation": "sum",
    })
    
    # Clip to window
    daily = daily.loc[CALIB_START:CALIB_END].dropna(subset=["Moisture_1.5m", "Moisture_3m"])
    if daily.empty:
        raise ValueError(f"No sensor data found between {CALIB_START} and {CALIB_END}!")

    print(f"  Sensor window: {daily.index[0].date()} to {daily.index[-1].date()}")
    print(f"  Days with data: {len(daily)}")
    return daily

def build_rainfall_time_series(daily):
    """
    Build weekly rainfall points for the GSI boundary function.
    X = elapsed seconds from simulation start (mid-week is best representative
        timestamp; using end-of-week is fine too but week 0 must be > 0s).
    Y = average in/day for the week.
    BUG FIX: week 0 previously mapped to elapsed_s=0 only when resampled.
             Ensure all X > 0 so GeoStudio does not treat them as t0 markers.
    """
    mm_to_in = 1.0 / 25.4
    weekly = daily["Precipitation"].resample("W").sum()
    points = []
    t0 = daily.index[0]
    for ts, total_mm in weekly.items():
        # ts is end-of-week; elapsed measured from calib window start
        elapsed_s = int((ts - t0).total_seconds())
        elapsed_s = max(elapsed_s, 86400)   # ensure X > 0
        rain_in_per_day = (total_mm * mm_to_in) / 7.0
        points.append({"X": float(elapsed_s), "Y": rain_in_per_day})
    return points

# ---------------------------------------------------------------------------
# SWCC Logic
# ---------------------------------------------------------------------------
def apply_swcc(pore_pressure_psf, swcc_pts, theta_sat):
    if pore_pressure_psf >= 0:
        return theta_sat
    suction = -pore_pressure_psf
    if suction <= swcc_pts[0][0]:
        return theta_sat
    if suction >= swcc_pts[-1][0]:
        return swcc_pts[-1][1]
    for i in range(len(swcc_pts) - 1):
        x0, y0 = swcc_pts[i]
        x1, y1 = swcc_pts[i + 1]
        if x0 <= suction <= x1:
            t = (suction - x0) / (x1 - x0)
            return y0 + t * (y1 - y0)
    return swcc_pts[-1][1]

# ---------------------------------------------------------------------------
# Objective
# ---------------------------------------------------------------------------
iter_count = [0]
trial_log  = []

# Node indices found once before optimisation - shared across all trials
_sensor_node_idx_shallow = None
_sensor_node_idx_deep    = None


def find_sensor_nodes(gsz_path: str) -> tuple[int | None, int | None]:
    """
    Read mesh node XY coordinates directly from the .gsz XML via PyGeoStudio,
    then return the indices of the nearest nodes to each sensor location.
    PyGeoStudio reads the raw XML geometry - no GSI API, no results query needed.
    The mesh never changes between trials so this is called exactly once.
    """
    import zipfile, xml.etree.ElementTree as ET

    gsz_stem = os.path.splitext(os.path.basename(gsz_path))[0]
    node_x, node_y = [], []

    try:
        with zipfile.ZipFile(gsz_path, 'r') as z:
            # Root XML is the one not inside an analysis subfolder
            analysis_folders = ["FS", "Slope Stability", "Initial Condition", "Rainfall Simulation"]
            xml_candidates = [
                name for name in z.namelist()
                if name.endswith(gsz_stem + ".xml")
                and not any(name.startswith(af + "/") for af in analysis_folders)
            ]
            if not xml_candidates:
                # Fallback: any root-level xml
                xml_candidates = [n for n in z.namelist()
                                  if n.endswith(".xml") and "/" not in n]
            if not xml_candidates:
                print("  ERROR: No root XML found in .gsz archive.")
                return None, None

            xml_bytes = z.read(xml_candidates[0])

        root = ET.fromstring(xml_bytes.decode("utf-8", errors="replace"))

        # GeoStudio XML stores mesh nodes as:
        # <MeshNodes Len="N"><Node X="..." Y="..." /> ...</MeshNodes>
        # or nested under <Mesh><Nodes>...
        for tag in ["MeshNodes", "Nodes"]:
            container = root.find(f".//{tag}")
            if container is not None:
                for node in container:
                    try:
                        node_x.append(float(node.get("X") or node.get("x")))
                        node_y.append(float(node.get("Y") or node.get("y")))
                    except (TypeError, ValueError):
                        continue
                if node_x:
                    print(f"  PyGeoStudio mesh read: {len(node_x)} nodes via <{tag}>")
                    break

        # Fallback: search all elements with X and Y attributes (broader sweep)
        if not node_x:
            seen = set()
            for elem in root.iter():
                x_attr = elem.get("X") or elem.get("x")
                y_attr = elem.get("Y") or elem.get("y")
                if x_attr is not None and y_attr is not None:
                    try:
                        x, y = float(x_attr), float(y_attr)
                        key = (round(x, 4), round(y, 4))
                        if key not in seen:
                            seen.add(key)
                            node_x.append(x)
                            node_y.append(y)
                    except ValueError:
                        continue
            if node_x:
                print(f"  PyGeoStudio mesh read: {len(node_x)} unique XY points (broad sweep)")

    except Exception as e:
        print(f"  ERROR reading mesh from GSZ: {e}")
        return None, None

    if not node_x:
        print("  ERROR: No node coordinates found in XML.")
        print("  -> Set SENSOR_NODE_SHALLOW and SENSOR_NODE_DEEP manually in CONFIG.")
        return None, None

    arr_x = np.array(node_x)
    arr_y = np.array(node_y)

    dist_s = np.sqrt((arr_x - CREST_X)**2 + (arr_y - SENSOR_Y_SHALLOW)**2)
    dist_d = np.sqrt((arr_x - CREST_X)**2 + (arr_y - SENSOR_Y_DEEP)**2)
    idx_s  = int(np.argmin(dist_s))
    idx_d  = int(np.argmin(dist_d))

    print(f"  Sensor node (shallow 1.5m): idx={idx_s}  "
          f"x={arr_x[idx_s]:.2f}  y={arr_y[idx_s]:.2f}  "
          f"dist={dist_s[idx_s]:.2f} ft")
    print(f"  Sensor node (deep 3m)     : idx={idx_d}  "
          f"x={arr_x[idx_d]:.2f}  y={arr_y[idx_d]:.2f}  "
          f"dist={dist_d[idx_d]:.2f} ft")

    if dist_s[idx_s] > 5.0 or dist_d[idx_d] > 5.0:
        print("  WARNING: Nearest node is >5 ft from sensor target coordinates.")
        print(f"    CREST_X={CREST_X}, SENSOR_Y_SHALLOW={SENSOR_Y_SHALLOW:.2f}, "
              f"SENSOR_Y_DEEP={SENSOR_Y_DEEP:.2f}")
        print("    Check that these coordinates are within the mesh boundary.")

    return idx_s, idx_d


def _set_material_prop(project, param, value, analysis=None):
    """Set a scalar material property. Defaults to SEEP_TRANSIENT analysis."""
    project.Set(gsi.SetRequest(
        analysis=analysis or SEEP_TRANSIENT,
        object=param,
        data=gsi.Value(number_value=float(value))
    ))

def _update_rainfall(project, rain_points):
    # 1. Update the Rainfall function dynamically by finding the climate fn named "rainfall"
    req = gsi.GetRequest(analysis=SEEP_TRANSIENT, object="Functions.Boundary.ClimateFns")
    res = project.Get(req)
    fn_list = res.data.list_value.values
    
    rainfall_idx = -1
    for i, item in enumerate(fn_list):
        if item.struct_value.fields.get('Name') and item.struct_value.fields['Name'].string_value == "rainfall":
            rainfall_idx = i
            break
            
    if rainfall_idx != -1:
        lv = ListValue()
        pt_list = []
        for pt in rain_points:
            s = Struct()
            s.fields["X"].number_value = pt["X"]
            s.fields["Y"].number_value = pt["Y"]
            v = Value()
            v.struct_value.CopyFrom(s)
            pt_list.append(v)
        lv.values.extend(pt_list)
        
        project.Set(gsi.SetRequest(
            analysis=SEEP_TRANSIENT,
            object=f"Functions.Boundary.ClimateFns[{rainfall_idx}].Points",
            data=gsi.Value(list_value=lv)
        ))
    else:
        print("  WARNING: Could not find 'rainfall' function via GSI!")

    # 2. Update TimeIncrements
    n_weeks = len(rain_points)
    start_s = 432000
    week_s = 7 * 86400
    dur_s = n_weeks * week_s
    
    ti_struct = Struct()
    ti_struct.fields["Start"].number_value = start_s
    ti_struct.fields["Duration"].number_value = dur_s
    ti_struct.fields["IncrementOption"].string_value = "Exponential"
    ti_struct.fields["IncrementCount"].string_value = str(n_weeks)
    
    ts_list = ListValue()
    for w in range(1, n_weeks + 1):
        s = Struct()
        s.fields["Step"].number_value = float(week_s)
        s.fields["ElapsedTime"].number_value = float(start_s + w * week_s)
        s.fields["Save"].bool_value = True
        v = Value()
        v.struct_value.CopyFrom(s)
        ts_list.values.append(v)
    
    ti_struct.fields["TimeSteps"].list_value.CopyFrom(ts_list)
    project.Set(gsi.SetRequest(
        analysis=SEEP_TRANSIENT,
        object="CurrentAnalysis.TimeIncrements",
        data=gsi.Value(struct_value=ti_struct)
    ))

def objective(params, swcc_pts, theta_sat, rain_points, obs_daily):
    log_ksat, log_kyx, log_ksat_uyc = params
    
    log_ksat     = np.clip(log_ksat,     LOG_KSAT_WYC_MIN, LOG_KSAT_WYC_MAX)
    log_kyx      = np.clip(log_kyx,      LOG_KYX_MIN,      LOG_KYX_MAX)
    log_ksat_uyc = np.clip(log_ksat_uyc, LOG_KSAT_UYC_MIN, LOG_KSAT_UYC_MAX)

    ksat_wyc  = 10 ** log_ksat
    kyx_awyc  = 10 ** log_kyx
    ksat_uyc  = 10 ** log_ksat_uyc

    iter_count[0] += 1
    idx = iter_count[0]
    print(f"  Trial {idx:3d}: KSat_WYC={ksat_wyc:.3e}  KYX_AWYC={kyx_awyc:.0f}  KSat_UYC={ksat_uyc:.3e}", end="", flush=True)

    try:
        shutil.copy2(BASE_GSZ, WORK_GSZ)
        project = gsi.OpenProject(WORK_GSZ)
        
        # Set Hyd properties
        _set_material_prop(project, 'Materials["Weathered Yazoo Clay"].Hydraulic.KSat', ksat_wyc)
        _set_material_prop(project, 'Materials["Weathered Yazoo Clay in Active Zone"].Hydraulic.KYXRatio', kyx_awyc)
        _set_material_prop(project, 'Materials["Unweathered Yazoo Clay"].Hydraulic.KSat', ksat_uyc)
        
        _update_rainfall(project, rain_points)

        # Solve
        solve_res = project.SolveAnalyses(gsi.SolveAnalysesRequest(analyses=[SEEP_TRANSIENT], solve_dependencies=True))
        if not solve_res.all_succeeded:
            print("  →  FAILED: Solver did not complete successfully.")
            return 9999.0

        project.LoadResults(gsi.LoadResultsRequest(analysis=SEEP_TRANSIENT))
        
        # Find how many output steps we have
        req_steps = project.QueryResultsAvailability(gsi.QueryResultsAvailabilityRequest(analysis=SEEP_TRANSIENT))
        # Usually step index starts at 1, step 0 is initial cond.
        avail_steps = req_steps.available_steps
        
        # Node indices resolved once before optimisation in main() via find_sensor_nodes()
        node_idx_shallow = _sensor_node_idx_shallow
        node_idx_deep    = _sensor_node_idx_deep
        if node_idx_shallow is None or node_idx_deep is None:
            print("  ->  WARNING: Sensor node indices not set.")
            return 9999.0

        sim_results = {}
        for step_idx in avail_steps:
            if step_idx <= 0: continue

            # Get elapsed time for this step
            t_req = gsi.QueryResultsRequest(
                analysis=SEEP_TRANSIENT, step=step_idx,
                table=gsi.ResultType.Time,
                dataparams=[gsi.DataParamType.eAbsoluteTime],
            )
            t_res = project.QueryResults(t_req)
            ev = t_res.results.get(gsi.DataParamType.eAbsoluteTime)
            if not ev or not ev.values: continue
            abs_time = ev.values[0]
            rel_s    = abs_time - 432000

            # Query pore water pressure at ALL nodes, then pick by index
            p_req = gsi.QueryResultsRequest(
                analysis=SEEP_TRANSIENT, step=step_idx,
                table=gsi.ResultType.Node,
                dataparams=[gsi.DataParamType.eWaterPressure],
            )
            try:
                p_res   = project.QueryResults(p_req)
                p_entry = p_res.results.get(gsi.DataParamType.eWaterPressure)
                if p_entry and len(p_entry.values) > max(node_idx_shallow, node_idx_deep):
                    pwp_s = p_entry.values[node_idx_shallow]
                    pwp_d = p_entry.values[node_idx_deep]
                    sim_results[rel_s] = {
                        "vwc_shallow": apply_swcc(pwp_s, swcc_pts, theta_sat),
                        "vwc_deep":    apply_swcc(pwp_d, swcc_pts, theta_sat),
                    }
            except Exception:
                pass

        if not sim_results:
            print("  →  WARNING: No sim results parsed.")
            return 9999.0

        # Calculate RMSE
        # sim_results keys are in seconds from calib start; obs_daily index is timestamps.
        # We match each OBSERVED daily reading to the NEAREST simulated weekly step.
        sim_keys = sorted(sim_results.keys())
        errors = []
        for ts, row in obs_daily.iterrows():
            elapsed_s = int((ts - obs_daily.index[0]).total_seconds())
            closest_t = min(sim_keys, key=lambda t: abs(t - elapsed_s))
            sim = sim_results[closest_t]
            obs_s = row["Moisture_1.5m"]
            obs_d = row["Moisture_3m"]
            if np.isfinite(obs_s) and np.isfinite(sim["vwc_shallow"]):
                errors.append((sim["vwc_shallow"] - obs_s)**2)
            if np.isfinite(obs_d) and np.isfinite(sim["vwc_deep"]):
                errors.append((sim["vwc_deep"] - obs_d)**2)

        rmse = np.sqrt(np.mean(errors)) if errors else 9999.0
        print(f"  →  RMSE = {rmse:.5f}")

        trial_log.append({
            "trial":        idx,
            "log_ksat_wyc": round(log_ksat, 4), "log_kyx": round(log_kyx, 4), "log_ksat_uyc": round(log_ksat_uyc, 4),
            "ksat_wyc": ksat_wyc, "kyx_awyc": kyx_awyc, "ksat_uyc": ksat_uyc, "rmse": rmse,
        })
        return rmse

    except Exception as e:
        print(f"  →  EXCEPTION: {e}")
        trial_log.append({
            "trial": idx, "log_ksat_wyc": round(log_ksat, 4),
            "log_kyx": round(log_kyx, 4), "log_ksat_uyc": round(log_ksat_uyc, 4),
            "ksat_wyc": ksat_wyc, "kyx_awyc": kyx_awyc, "ksat_uyc": ksat_uyc,
            "rmse": 9999.0,
        })
        return 9999.0
    finally:
        if 'project' in locals() and project is not None:
            try: project.Close()
            except: pass
        if os.path.exists(WORK_GSZ):
            try: os.remove(WORK_GSZ)
            except: pass

SWCC_XLSX      = os.path.join(CALIB_DIR, "SWCC_Models_rev1 4.xlsm")

def get_base_swcc():
    """
    Read the SWCC curve for the AWYC material from the Excel workbook.
    The workbook contains columns: Suction (psf) | Vol. WC | ...
    We look for the sheet named 'AWYC' (or containing 'AWYC').
    Falls back to the GSI API path if Excel is unavailable.
    """
    import openpyxl

    if os.path.exists(SWCC_XLSX):
        try:
            wb = openpyxl.load_workbook(SWCC_XLSX, read_only=True, data_only=True)
            # Look for a sheet with AWYC in its name
            target_ws = None
            for name in wb.sheetnames:
                if "AWYC" in name.upper():
                    target_ws = wb[name]
                    break
            if target_ws is None:
                # Fall back to first sheet
                target_ws = wb.active

            swcc = []
            theta_sat = None
            for row in target_ws.iter_rows(min_row=2, values_only=True):
                try:
                    suction = float(row[0])   # psf
                    vwc     = float(row[1])   # vol. water content
                    swcc.append((suction, vwc))
                    if theta_sat is None or vwc > theta_sat:
                        theta_sat = vwc       # sat = max VWC in curve
                except (TypeError, ValueError, IndexError):
                    continue

            if swcc:
                print(f"  SWCC loaded from Excel: {len(swcc)} points, θsat={theta_sat:.3f}")
                return sorted(swcc, key=lambda x: x[0]), theta_sat
        except Exception as e:
            print(f"  WARNING: could not read SWCC Excel: {e}. Falling back to API.")

    # Fallback: read via GSI API (Functions.Material.Hydraulic.VolWCFns)
    try:
        project = gsi.OpenProject(BASE_GSZ)
        try:
            req = gsi.GetRequest(analysis=SEEP_TRANSIENT, object="Functions")
            res = project.Get(req)
            d   = MessageToDict(res.data.struct_value)
            fns = d.get("Material", {}).get("Hydraulic", {}).get("VolWCFns", [])
            for fn in fns:
                if "AWYC" in fn.get("Name", ""):
                    pts = fn.get("Points", [])
                    swcc = [(float(p["Value"][0]), float(p["Value"][1])) for p in pts if len(p.get("Value", [])) >= 2]
                    t_sat = fn.get("Estimate", {}).get("SatWC", {}).get("Value", 0.55)
                    return sorted(swcc, key=lambda x: x[0]), float(t_sat)
        finally:
            project.Close()
    except Exception as e:
        print(f"  ERROR: SWCC API fallback also failed: {e}")

    # Hard-coded fallback from probe_out3.txt (VOL WC AWYC, θsat=0.55)
    print("  Using hard-coded AWYC SWCC fallback.")
    AWYC_SWCC_HCF = [
        (0.0, 0.55), (0.209, 0.5500), (0.383, 0.5500), (0.702, 0.5500),
        (1.286, 0.5500), (2.358, 0.5500), (4.321, 0.5500), (7.921, 0.5499),
        (14.519, 0.5498), (26.614, 0.5494), (48.783, 0.5486), (89.417, 0.5466),
        (163.9, 0.5419), (300.43, 0.5308), (550.68, 0.5066), (1009.38, 0.4604),
        (1850.17, 0.3907), (3391.33, 0.3128), (6216.24, 0.2460),
        (11394.25, 0.1962), (20885.43, 0.1599),
    ]
    return AWYC_SWCC_HCF, 0.55


def main():
    print("=" * 60)
    print("SEEP/W Hydraulic Calibration (GSI API)")
    print(f"  Cal window   : {CALIB_START} to {CALIB_END}")
    print("=" * 60)
    
    if not os.path.exists(BASE_GSZ) or not os.path.exists(SENSOR_CSV):
        print("Missing required base files!")
        return

    obs_daily = load_sensor_data()
    rain_points = build_rainfall_time_series(obs_daily)
    
    swcc_pts, theta_sat = get_base_swcc()
    if not swcc_pts:
        print("Error getting SWCC")
        return

    # --- Locate mesh nodes nearest to sensors (done once, shared across all trials) ---
    global _sensor_node_idx_shallow, _sensor_node_idx_deep
    if SENSOR_NODE_SHALLOW is not None and SENSOR_NODE_DEEP is not None:
        _sensor_node_idx_shallow = SENSOR_NODE_SHALLOW
        _sensor_node_idx_deep    = SENSOR_NODE_DEEP
        print(f"\nUsing manual node overrides: shallow={SENSOR_NODE_SHALLOW}  deep={SENSOR_NODE_DEEP}")
    else:
        print("\nLocating mesh nodes nearest to sensors...")
        _sensor_node_idx_shallow, _sensor_node_idx_deep = find_sensor_nodes(BASE_GSZ)
        if _sensor_node_idx_shallow is None:
            print("\nERROR: Could not find sensor nodes. Set SENSOR_NODE_SHALLOW and "
                  "SENSOR_NODE_DEEP manually in CONFIG and re-run.")
            return

    x0 = np.array([
        np.log10(5.640e-06),      # KSat_WYC
        np.log10(80827.0),        # KYX_AWYC
        np.log10(KSAT_UYC_BASE),  # KSat_UYC
    ])

    result = minimize(
        objective, x0, args=(swcc_pts, theta_sat, rain_points, obs_daily),
        method="Nelder-Mead", options={"maxiter": MAX_OPT_ITER, "xatol": 0.05, "fatol": 0.001}
    )

    print("\nOptimisation complete")
    print(f"  Status    : {result.message}")
    print(f"  Best RMSE : {result.fun:.5f}")

    if trial_log:
        pd.DataFrame(trial_log).to_csv(OUT_CSV, index=False)
        print(f"Trial log saved to {OUT_CSV}")

    # Final run to save OUT_GSZ
    try:
        shutil.copy2(BASE_GSZ, OUT_GSZ)
        project = gsi.OpenProject(OUT_GSZ)
        best_ksat, best_kyx, best_uyc = 10**result.x[0], 10**result.x[1], 10**result.x[2]

        # Apply calibrated Ksat to BOTH Initial Condition and Transient analyses
        # so that the downstream SLOPE/W calibration inherits them correctly.
        for analysis in [SEEP_INITIAL, SEEP_TRANSIENT]:
            try:
                _set_material_prop(project, 'Materials["Weathered Yazoo Clay"].Hydraulic.KSat',
                                   best_ksat, analysis=analysis)
                _set_material_prop(project, 'Materials["Weathered Yazoo Clay in Active Zone"].Hydraulic.KYXRatio',
                                   best_kyx, analysis=analysis)
                _set_material_prop(project, 'Materials["Unweathered Yazoo Clay"].Hydraulic.KSat',
                                   best_uyc, analysis=analysis)
            except Exception:
                pass     # some props may not exist on Initial Condition

        _update_rainfall(project, rain_points)
        project.SolveAnalyses(gsi.SolveAnalysesRequest(analyses=[SEEP_TRANSIENT], solve_dependencies=True))
        print(f"Saved optimized GSZ to {OUT_GSZ}")
    finally:
        project.Close()

if __name__ == "__main__":
    main()
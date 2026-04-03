"""
SLOPE/W Calibration with Inclinometer Validation
==================================================
Uses the GeoStudio official Python scripting API (gsi — no ZIP/XML manipulation).

Requires
--------
- Python 3.12.x  (the gsi wheel requires Python 3.12)
- gsi installed:
    pip install -r "C:\\Program Files\\Seequent\\GeoStudio 2025.2\\API\\requirements.txt"
    pip install openpyxl scipy

Base GSZ requirements
---------------------
The BASE_GSZ must be:
  1. Solved with calibrated SEEP/W parameters (KSat_WYC, KYX_AWYC from calibrate_seep.py)
  2. Rainfall Simulation configured for the Oct–Nov 2020 calibration event
  3. FS analysis linked to the last timestep of the Rainfall Simulation

Recommended base: Metro-Center-calibrated-slope.gsz (open in GeoStudio, configure
  Oct–Nov 2020 rainfall under the Rainfall Simulation BC, and solve once first).

Dual objective
--------------
  cost = W_FS * (FS - 1.0)^2 + W_INCL * depth_error^2

  depth_error: normalised mismatch between the critical slip surface depth
               at the borehole location and the inclinometer-observed shear zone depth.
               If the slip circle doesn't reach the borehole X, the penalty uses
               the lateral gap instead (drives the optimizer toward deeper, wider circles).

Inclinometer data (S2_Inclinometer_Y1/Y2.xlsx)
-----------------------------------------------
  - A-direction (downslope) cumulative profile change in inches at 2-ft depth intervals
  - Shear zone depth = deepest depth (from surface) where cumulative displacement
    still exceeds DISP_THRESHOLD at the failure-date reading
  - Y1 goes to 27 ft (deeper), Y2 goes to 21 ft
  - Nov 3 2020 reading is used (closest available date to the Nov 2020 failure)

Usage
-----
    C:\\Python312\\python.exe calibrate_slope_incl.py
"""

from __future__ import annotations
import os, sys, csv, shutil, datetime, warnings
import numpy as np
import openpyxl
from scipy.optimize import minimize

warnings.filterwarnings("ignore")

# ---------------------------------------------------------------------------
# gsi import
# ---------------------------------------------------------------------------
try:
    import grpc                                          # noqa: F401
    from google.protobuf.json_format import MessageToDict  # noqa: F401
    import gsi
    _GSI_OK = True
except ImportError as _e:
    _GSI_OK = False
    _GSI_ERROR = str(_e)

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------

CALIB_DIR = r"E:\Github\MCP_Geostudio\calibration"

# Base file: must have calibrated SEEP/W params + Oct-Nov 2020 rainfall.
# The script opens this file, modifies c/phi via API, and solves in place.
# A working copy (WORK_GSZ) is used so BASE_GSZ is never modified.
BASE_GSZ  = os.path.join(CALIB_DIR, "Metro-Center-calibrated-slope.gsz")
WORK_GSZ  = os.path.join(CALIB_DIR, "_slope_incl_work.gsz")     # working copy
OUT_GSZ   = os.path.join(CALIB_DIR, "Metro-Center-slope-final.gsz")
LOG_CSV   = os.path.join(CALIB_DIR, "slope_incl_trial_log.csv")

INCL_Y1   = os.path.join(CALIB_DIR, "S2_Inclinometer_Y1.xlsx")
INCL_Y2   = os.path.join(CALIB_DIR, "S2_Inclinometer_Y2.xlsx")

# Analysis names
FS_ANALYSIS = "FS"
FS_STEP     = 1   # FS analysis has a single save step

# Borehole / model geometry  (S2 sensor, Crest position)
BOREHOLE_X   = 195.0   # ft — mid-crest platform (X = 179–220)
CREST_Y      = 83.0    # ft — crest surface elevation
FAILURE_DATE = datetime.datetime(2020, 11, 3)   # closest inclinometer reading
DISP_THRESH  = 0.005   # in — below this = no movement = above shear zone

# Baseline parameters (forensic study values)
C_AWYC_BASE,  PHI_AWYC_BASE = 79.3,  19.0
C_WYC_BASE,   PHI_WYC_BASE  = 248.5, 19.0

# Bounds
C_AWYC_MIN,  C_AWYC_MAX  =   0.0, 250.0
PHI_AWYC_MIN,PHI_AWYC_MAX=  10.0,  25.0
C_WYC_MIN,   C_WYC_MAX   =  50.0, 500.0
PHI_WYC_MIN, PHI_WYC_MAX =  12.0,  25.0

# Objective weights
FS_TARGET = 1.0
W_FS      = 1.0   # weight on (FS - 1.0)^2
W_INCL    = 0.5   # weight on inclinometer depth mismatch

MAX_OPT_ITER = 60

# Materials to update in every analysis (materials are shared across analyses)
ANALYSES_ALL = ["Initial Condition", "Rainfall Simulation", "Slope Stability", "FS"]

# ---------------------------------------------------------------------------
# Inclinometer parsing
# ---------------------------------------------------------------------------

def _parse_incl_block(ws, header_row: int, data_start_row: int) -> dict:
    """
    Parse one block of depth–displacement data from an inclinometer worksheet.

    Returns {depth_ft: {date: displacement_in, ...}, ...}
    Depths are in the column at index 2 (0-based), dates start at index 3.
    """
    row5 = [cell.value for cell in ws[header_row]]
    date_col_indices = [
        (j, v) for j, v in enumerate(row5)
        if isinstance(v, datetime.datetime)
    ]
    if not date_col_indices:
        return {}

    data: dict[float, dict] = {}
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        depth = row[2]
        if depth is None or not isinstance(depth, (int, float)):
            break
        depth = float(depth)
        data[depth] = {}
        for col_idx, dt in date_col_indices:
            val = row[col_idx]
            if val is not None:
                try:
                    data[depth][dt] = float(val)
                except (TypeError, ValueError):
                    pass
    return data


def load_inclinometer() -> float | None:
    """
    Parse both inclinometer Excel files and return the shear zone depth (ft)
    at the failure date.

    Method:
      1. Read the A-direction block (first data block, header at row 5) from each file.
      2. Find the reading closest to FAILURE_DATE.
      3. Shear zone depth = deepest depth where cumulative displacement > DISP_THRESH.
      4. Average the estimate from both files (Y1 is used preferentially as it goes deeper).

    Returns: obs_depth_ft (float) — depth of shear zone below CREST_Y,
             or None if data cannot be parsed.
    """
    shear_depths: list[float] = []

    for fpath, label in [(INCL_Y1, "Y1"), (INCL_Y2, "Y2")]:
        if not os.path.exists(fpath):
            print(f"  WARNING: inclinometer file not found: {fpath}")
            continue

        wb = openpyxl.load_workbook(fpath, data_only=True)
        ws = wb.active

        # A-direction block: header at row 5, data starts at row 7
        # (row 6 is a metadata row with [-1, 2, 1, ...])
        data = _parse_incl_block(ws, header_row=5, data_start_row=7)
        if not data:
            print(f"  WARNING: could not parse inclinometer block in {label}")
            continue

        # Find reading closest to FAILURE_DATE
        all_dates: set[datetime.datetime] = set()
        for depth_data in data.values():
            all_dates.update(depth_data.keys())
        if not all_dates:
            continue

        closest_dt = min(all_dates, key=lambda d: abs((d - FAILURE_DATE).days))
        days_off   = abs((closest_dt - FAILURE_DATE).days)
        print(f"  {label}: using reading {str(closest_dt)[:10]}  "
              f"({days_off} days from {FAILURE_DATE.date()})")

        # Build depth–displacement profile at closest date
        profile: list[tuple[float, float]] = []
        for depth in sorted(data.keys()):
            disp = data[depth].get(closest_dt)
            if disp is not None:
                profile.append((depth, abs(disp)))

        if not profile:
            continue

        # Shear zone: deepest depth where displacement > threshold (from bottom up)
        profile_sorted = sorted(profile, key=lambda x: x[0], reverse=True)  # deep → shallow
        shear_depth = None
        for depth, disp in profile_sorted:
            if disp > DISP_THRESH:
                shear_depth = depth
                break  # deepest depth with non-trivial movement

        if shear_depth is None:
            shear_depth = max(d for d, _ in profile)  # all movement → take max depth

        print(f"  {label}: shear zone depth = {shear_depth:.0f} ft  "
              f"(SHEAR_Y ≈ {CREST_Y - shear_depth:.1f} ft in model coords)")

        # Show displacement profile (brief)
        for depth, disp in sorted(profile):
            marker = " ← shear zone" if depth == shear_depth else ""
            print(f"    {depth:4.0f} ft  {disp:.4f} in{marker}")

        shear_depths.append(shear_depth)

    if not shear_depths:
        return None

    # Prefer Y1 (deeper) if available; otherwise average
    obs_depth = max(shear_depths)   # use the deeper/more conservative estimate
    print(f"\n  Observed shear zone depth : {obs_depth:.1f} ft "
          f"(SHEAR_Y = {CREST_Y - obs_depth:.1f} ft)")
    return obs_depth

# ---------------------------------------------------------------------------
# gsi helpers
# ---------------------------------------------------------------------------

def _require_gsi():
    if not _GSI_OK:
        raise RuntimeError(
            f"gsi not available: {_GSI_ERROR}\n"
            "Install: pip install -r "
            r'"C:\Program Files\Seequent\GeoStudio 2025.2\API\requirements.txt"'
            "\nRequires Python 3.12.x"
        )


def _set_strength(project, analysis: str, c_awyc: float, phi_awyc: float,
                  c_wyc: float, phi_wyc: float):
    """Set c′ and φ′ for AWYC and WYC in the specified analysis."""
    updates = [
        ("Weathered Yazoo Clay in Active Zone", "CohesionPrime", c_awyc),
        ("Weathered Yazoo Clay in Active Zone", "PhiPrime",      phi_awyc),
        ("Weathered Yazoo Clay",                "CohesionPrime", c_wyc),
        ("Weathered Yazoo Clay",                "PhiPrime",      phi_wyc),
    ]
    for mat, prop, val in updates:
        project.Set(gsi.SetRequest(
            analysis=analysis,
            object=f'Materials["{mat}"].{prop}',
            data=gsi.Value(number_value=float(val)),
        ))


def _get_critical_fs(project) -> float | None:
    """Query minimum FS from the critical slip surface."""
    project.LoadResults(gsi.LoadResultsRequest(analysis=FS_ANALYSIS))
    resp = project.QueryResults(gsi.QueryResultsRequest(
        analysis=FS_ANALYSIS,
        step=FS_STEP,
        table=gsi.ResultType.CriticalSlip,
        dataparams=[gsi.DataParamType.eSlipFOSMin],
    ))
    entry = resp.results.get(gsi.DataParamType.eSlipFOSMin)
    if entry and entry.values:
        vals = [v for v in entry.values if 0.1 < v < 100]
        return min(vals) if vals else None
    return None


def _get_slip_depth_at_borehole(project) -> tuple[float | None, dict]:
    """
    Compute depth of critical slip surface at BOREHOLE_X using the slip circle
    center and radius.

    Returns (depth_ft, info_dict).
    depth_ft = CREST_Y - slip_Y  if borehole is inside the circle, else None.
    info_dict contains geometry for logging.
    """
    # The critical slip surface geometry is in ResultType.Slip (all surfaces)
    # We read SlipCenterX, SlipCenterY, SlipRadiusX for the row with min FS
    resp = project.QueryResults(gsi.QueryResultsRequest(
        analysis=FS_ANALYSIS,
        step=FS_STEP,
        table=gsi.ResultType.Slip,
        dataparams=[
            gsi.DataParamType.eSlipFOSMin,
            gsi.DataParamType.eXCoord,     # SlipCenterX equivalent
            gsi.DataParamType.eYCoord,     # SlipCenterY equivalent
        ],
    ))

    fos_entry = resp.results.get(gsi.DataParamType.eSlipFOSMin)
    cx_entry  = resp.results.get(gsi.DataParamType.eXCoord)
    cy_entry  = resp.results.get(gsi.DataParamType.eYCoord)

    if not all([fos_entry, cx_entry, cy_entry]):
        return None, {}

    fos_vals = list(fos_entry.values)
    cx_vals  = list(cx_entry.values)
    cy_vals  = list(cy_entry.values)

    if not fos_vals:
        return None, {}

    # Find index of minimum valid FS
    valid = [(i, f) for i, f in enumerate(fos_vals) if 0.1 < f < 100]
    if not valid:
        return None, {}

    best_idx = min(valid, key=lambda x: x[1])[0]
    cx = cx_vals[best_idx] if best_idx < len(cx_vals) else None
    cy = cy_vals[best_idx] if best_idx < len(cy_vals) else None

    if cx is None or cy is None:
        return None, {}

    # Use CriticalSlip table for radius — query eXCoord to get column X positions
    # and infer radius from the extent
    col_resp = project.QueryResults(gsi.QueryResultsRequest(
        analysis=FS_ANALYSIS,
        step=FS_STEP,
        table=gsi.ResultType.Column,
        dataparams=[gsi.DataParamType.eXCoord, gsi.DataParamType.eYCoord],
    ))
    col_x = list(col_resp.results.get(gsi.DataParamType.eXCoord, type('', (), {'values': []})()).values)
    col_y = list(col_resp.results.get(gsi.DataParamType.eYCoord, type('', (), {'values': []})()).values)

    # Infer radius from column geometry (distance from center to any column base)
    radius = None
    if col_x and col_y:
        dists = [np.sqrt((x - cx)**2 + (y - cy)**2) for x, y in zip(col_x, col_y)]
        radius = np.mean(dists)  # should all be ~equal for a circular arc

    info = {"cx": round(cx, 2), "cy": round(cy, 2), "radius": round(radius, 2) if radius else None}

    if radius is None:
        return None, info

    # Check if borehole X is within circle
    dx = BOREHOLE_X - cx
    if abs(dx) > radius:
        # Borehole is outside slip circle: return gap as negative depth (for penalty)
        gap = abs(dx) - radius
        info["outside_circle"] = True
        info["lateral_gap_ft"] = round(gap, 2)
        return None, info

    # Borehole is inside: compute slip surface Y at borehole X
    slip_y = cy - np.sqrt(radius**2 - dx**2)
    depth  = CREST_Y - slip_y
    info["slip_y_at_borehole"] = round(slip_y, 2)
    info["depth_at_borehole"]  = round(depth, 2)
    return max(0.0, depth), info

# ---------------------------------------------------------------------------
# Objective function
# ---------------------------------------------------------------------------

_iter_count = [0]
_trial_log: list[dict] = []


def objective(params: np.ndarray, obs_depth: float | None) -> float:
    c_awyc, phi_awyc, c_wyc, phi_wyc = (
        float(np.clip(params[0], C_AWYC_MIN,   C_AWYC_MAX)),
        float(np.clip(params[1], PHI_AWYC_MIN, PHI_AWYC_MAX)),
        float(np.clip(params[2], C_WYC_MIN,    C_WYC_MAX)),
        float(np.clip(params[3], PHI_WYC_MIN,  PHI_WYC_MAX)),
    )

    _iter_count[0] += 1
    idx = _iter_count[0]
    print(f"  Trial {idx:3d}: "
          f"c_AWYC={c_awyc:6.1f}  φ_AWYC={phi_awyc:5.2f}°  "
          f"c_WYC={c_wyc:6.1f}  φ_WYC={phi_wyc:5.2f}°", end="", flush=True)

    project = None
    try:
        project = gsi.OpenProject(WORK_GSZ)

        # Set strength in all analyses (shared materials)
        for analysis in ANALYSES_ALL:
            try:
                _set_strength(project, analysis, c_awyc, phi_awyc, c_wyc, phi_wyc)
            except Exception:
                pass  # analysis may not exist in all files

        # Solve: re-runs Initial Condition → Rainfall Simulation → FS
        project.SolveAnalyses(gsi.SolveAnalysesRequest(
            analyses=[FS_ANALYSIS],
            solve_dependencies=True,
        ))

        # Read FS
        fs = _get_critical_fs(project)

        # Read slip depth at borehole
        slip_depth, slip_info = _get_slip_depth_at_borehole(project)

        # FS term
        if fs is None:
            fs_cost = 9999.0
        else:
            fs_cost = W_FS * (fs - FS_TARGET) ** 2

        # Inclinometer depth term
        incl_cost = 0.0
        if obs_depth is not None:
            if slip_depth is not None:
                # Borehole inside circle: normalised depth mismatch
                incl_cost = W_INCL * ((slip_depth - obs_depth) / max(obs_depth, 1.0)) ** 2
            elif "lateral_gap_ft" in slip_info:
                # Borehole outside circle: penalise lateral gap
                gap_normalised = slip_info["lateral_gap_ft"] / max(BOREHOLE_X, 1.0)
                incl_cost = W_INCL * gap_normalised ** 2

        cost = fs_cost + incl_cost

        fs_str    = f"{fs:.4f}" if fs is not None else "N/A"
        depth_str = (f"{slip_depth:.1f} ft" if slip_depth is not None
                     else f"outside (gap={slip_info.get('lateral_gap_ft','?')} ft)")
        print(f"  →  FS={fs_str}  slip_depth@borehole={depth_str}  cost={cost:.5f}")

        _trial_log.append({
            "trial":     idx,
            "c_awyc":    round(c_awyc,   2),
            "phi_awyc":  round(phi_awyc, 3),
            "c_wyc":     round(c_wyc,    2),
            "phi_wyc":   round(phi_wyc,  3),
            "fs":        round(fs, 5) if fs else 9999,
            "slip_depth_ft":  round(slip_depth, 2) if slip_depth else None,
            "obs_depth_ft":   round(obs_depth, 1)  if obs_depth else None,
            "slip_cx":   slip_info.get("cx"),
            "slip_cy":   slip_info.get("cy"),
            "slip_r":    slip_info.get("radius"),
            "fs_cost":   round(fs_cost,   6),
            "incl_cost": round(incl_cost, 6),
            "total_cost":round(cost,      6),
        })
        return cost

    except Exception as e:
        print(f"  →  FAILED: {e}")
        _trial_log.append({
            "trial": idx, "c_awyc": c_awyc, "phi_awyc": phi_awyc,
            "c_wyc": c_wyc, "phi_wyc": phi_wyc,
            "fs": 9999, "slip_depth_ft": None, "obs_depth_ft": obs_depth,
            "slip_cx": None, "slip_cy": None, "slip_r": None,
            "fs_cost": 9999, "incl_cost": 0, "total_cost": 9999,
        })
        return 9999.0

    finally:
        if project is not None:
            project.Close()

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 70)
    print("SLOPE/W Calibration with Inclinometer Validation (gsi API)")
    print(f"  Base GSZ    : {os.path.basename(BASE_GSZ)}")
    print(f"  Failure date: {FAILURE_DATE.date()}")
    print(f"  Borehole    : X = {BOREHOLE_X} ft  (crest platform)")
    print(f"  FS target   : {FS_TARGET}")
    print(f"  Weights     : W_FS={W_FS}  W_INCL={W_INCL}")
    print(f"  Max trials  : {MAX_OPT_ITER}")
    print("=" * 70)

    # ---- Check gsi ----
    _require_gsi()

    # ---- Check files ----
    for p, lbl in [(BASE_GSZ, "Base GSZ"), (INCL_Y1, "Incl Y1"), (INCL_Y2, "Incl Y2")]:
        if not os.path.exists(p):
            print(f"\nERROR: {lbl} not found: {p}")
            sys.exit(1)

    # ---- Parse inclinometer ----
    print("\nLoading inclinometer data...")
    obs_depth = load_inclinometer()

    if obs_depth is None:
        print("\nWARNING: could not determine shear zone depth. "
              "Running with FS-only objective (W_INCL overridden to 0).")

    # ---- Prepare working copy of BASE_GSZ ----
    print(f"\nPreparing working copy → {os.path.basename(WORK_GSZ)}")
    shutil.copy2(BASE_GSZ, WORK_GSZ)

    # ---- Initial parameters ----
    x0 = np.array([C_AWYC_BASE, PHI_AWYC_BASE, C_WYC_BASE, PHI_WYC_BASE])
    print(f"\nStarting parameters (forensic study baseline):")
    print(f"  c_AWYC   = {x0[0]:.1f} psf   bounds [{C_AWYC_MIN}, {C_AWYC_MAX}]")
    print(f"  φ_AWYC   = {x0[1]:.2f}°     bounds [{PHI_AWYC_MIN}, {PHI_AWYC_MAX}]")
    print(f"  c_WYC    = {x0[2]:.1f} psf   bounds [{C_WYC_MIN}, {C_WYC_MAX}]")
    print(f"  φ_WYC    = {x0[3]:.2f}°     bounds [{PHI_WYC_MIN}, {PHI_WYC_MAX}]")
    if obs_depth is not None:
        print(f"\n  Inclinometer constraint: slip depth at X={BOREHOLE_X} → "
              f"target = {obs_depth:.1f} ft (SHEAR_Y ≈ {CREST_Y - obs_depth:.1f} ft)")

    print(f"\nRunning optimisation (max {MAX_OPT_ITER} trials)...\n")

    result = minimize(
        objective,
        x0,
        args=(obs_depth,),
        method="Nelder-Mead",
        options={
            "maxiter": MAX_OPT_ITER,
            "xatol":   0.5,     # 0.5 psf / 0.5 degrees
            "fatol":   1e-4,
            "disp":    True,
        },
    )

    # ---- Pick best trial from log ----
    valid = [t for t in _trial_log if isinstance(t["fs"], float) and t["fs"] < 900]
    if valid:
        best = min(valid, key=lambda t: t["total_cost"])
    else:
        best = {"c_awyc": result.x[0], "phi_awyc": result.x[1],
                "c_wyc": result.x[2], "phi_wyc": result.x[3],
                "fs": None, "slip_depth_ft": None}

    # ---- Summary ----
    print(f"\n{'=' * 70}")
    print("Optimisation complete")
    print(f"  Status  : {result.message}   Trials: {result.nit}")
    if best["fs"] is not None:
        print(f"  Best FS : {best['fs']:.4f}  |FS-1.0| = {abs(best['fs'] - FS_TARGET):.4f}")
    if best["slip_depth_ft"] is not None and obs_depth is not None:
        print(f"  Slip depth at borehole : {best['slip_depth_ft']:.1f} ft  "
              f"(observed {obs_depth:.1f} ft, error {best['slip_depth_ft'] - obs_depth:+.1f} ft)")
    print()
    print(f"  {'Parameter':<22} {'Baseline':>12}  {'Calibrated':>12}  {'Change':>10}")
    print(f"  {'-'*62}")
    for lbl, base, cal in [
        ("c_AWYC (psf)",   C_AWYC_BASE,  best["c_awyc"]),
        ("φ_AWYC (°)",    PHI_AWYC_BASE, best["phi_awyc"]),
        ("c_WYC (psf)",    C_WYC_BASE,   best["c_wyc"]),
        ("φ_WYC (°)",     PHI_WYC_BASE,  best["phi_wyc"]),
    ]:
        print(f"  {lbl:<22} {base:>12.2f}  {cal:>12.2f}  {cal - base:>+10.2f}")

    # ---- Save trial log ----
    if _trial_log:
        with open(LOG_CSV, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=_trial_log[0].keys())
            writer.writeheader()
            writer.writerows(_trial_log)
        print(f"\nTrial log → {LOG_CSV}")

    # ---- Final solve with best params → save as OUT_GSZ ----
    print(f"\nRunning final solve with best parameters → {os.path.basename(OUT_GSZ)} ...")
    shutil.copy2(BASE_GSZ, OUT_GSZ)
    project = gsi.OpenProject(OUT_GSZ)
    try:
        for analysis in ANALYSES_ALL:
            try:
                _set_strength(project, analysis,
                              best["c_awyc"], best["phi_awyc"],
                              best["c_wyc"],  best["phi_wyc"])
            except Exception:
                pass
        project.SolveAnalyses(gsi.SolveAnalysesRequest(
            analyses=[FS_ANALYSIS], solve_dependencies=True,
        ))
        final_fs = _get_critical_fs(project)
        print(f"  Final FS = {final_fs:.4f}" if final_fs else "  FS unreadable")
    finally:
        project.Close()

    # ---- Cleanup ----
    if os.path.exists(WORK_GSZ):
        os.remove(WORK_GSZ)

    print(f"\nDone.  Final GSZ → {OUT_GSZ}")


if __name__ == "__main__":
    main()
